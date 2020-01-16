from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .forms import FileUploadForm
from django.contrib import messages
from .models import FileData
import openpyxl
#from pyexcel_xls import get_data as xls_get
#from pyexcel_xlsx import get_data as xlsx_get
#from django.utils.datastructures import MultiValueDictKeyError
#from rest_framework.views import APIView



@login_required
def dashboard(request):
    return render(request,'app/index.html')


@login_required
def upload(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)

        if form.is_valid():
            file = request.FILES['file']

            workbook = openpyxl.load_workbook(file)
            first_sheet = workbook.get_sheet_names()[0]
            worksheet = workbook.get_sheet_by_name(first_sheet)

            last_row=worksheet.max_row

            data = []

            for row in worksheet.iter_rows(min_row=2, max_row=last_row): #Offset for header
                stock = FileData()
                #stock.user = request.user
                stock.account_number = row[0].value
                stock.date = row[1].value
                stock.details = row[2].value
                stock.withdrawal_amt = row[3].value
                stock.deposit_amt = row[4].value
                stock.balance = row[5].value
                #stock.date = datetime.today()
                #stock.confirm = 'approved'
                #stock.seller = row[7].value
                data.append(stock)
            # Bulk create datas
            FileData.objects.bulk_create(data)

            messages.success(request, f'File Uploaded!')
            return render(request,'app/upload.html',{'form': form})
    else:
        form = FileUploadForm()
        return render(request,'app/upload.html',{'form': form})
